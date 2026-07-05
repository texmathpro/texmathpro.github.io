import os
import base64
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

# -----------------------------------
# Configuration
# -----------------------------------

EXCEL_FILE = "HugoBlog.xlsx"

OUTPUT_FOLDER = "output"

REPO_OWNER = "texmathpro"
REPO_NAME = "texmathpro.github.io"
BRANCH = "main"
CONTENT_PATH = "content/posts/"

GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")

HEADERS = {
    "Authorization": f"token {GITHUB_TOKEN}",
    "Accept": "application/vnd.github+json"
}


# -----------------------------------
# Helpers
# -----------------------------------

def as_bool(value):
    if isinstance(value, bool):
        return value

    if value is None:
        return False

    return str(value).strip().lower() == "true"


def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return ""


# -----------------------------------
# Convert spreadsheet to markdown
# -----------------------------------

def convert_sheet_to_markdown():
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    markdown_files = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        if not row[0]:
            continue

        product_id = str(row[0]).strip()
        title = str(row[1] or "").strip()
        photo = str(row[2] or "").strip()

        body_regular = row[3] or ""
        body_discounted = row[4] or ""

        price_regular = float(row[5] or 0)
        price_discounted = float(row[6] or 0)

        is_discounted = as_bool(row[7])
        discount_until = format_date(row[8])

        is_published = as_bool(row[9])

        if not is_published:
            continue

        date = format_date(row[10])

        body = body_discounted if is_discounted else body_regular

        escaped_title = title.replace('"', '\\"')

        frontmatter = f"""---
title: "{escaped_title}"
date: {date}
productId: "{product_id}"
photo: "{photo}"
priceRegular: {price_regular}
priceDiscounted: {price_discounted}
isDiscounted: {str(is_discounted).lower()}
discountedUntil: "{discount_until}"
draft: false
---

"""

        markdown = frontmatter + str(body)

        markdown_files.append({
            "filename": f"{product_id}.md",
            "content": markdown
        })

    print(f"Generated {len(markdown_files)} markdown files.")

    return markdown_files


# -----------------------------------
# Save locally
# -----------------------------------

def save_markdown_files():
    files = convert_sheet_to_markdown()

    if not files:
        print("No published files.")
        return

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    for file in files:
        path = os.path.join(OUTPUT_FOLDER, file["filename"])

        with open(path, "w", encoding="utf-8") as f:
            f.write(file["content"])

    print(f"Saved {len(files)} markdown files to '{OUTPUT_FOLDER}'.")


# -----------------------------------
# Push to GitHub (single commit)
# -----------------------------------

def github_request(method, url, **kwargs):
    r = requests.request(method, url, headers=HEADERS, **kwargs)

    if not r.ok:
        raise Exception(f"{r.status_code}: {r.text}")

    return r.json()


def push_to_github():

    files = convert_sheet_to_markdown()

    if not files:
        print("No published files.")
        return

    # Step 1
    ref = github_request(
        "GET",
        f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/git/ref/heads/{BRANCH}"
    )

    latest_commit_sha = ref["object"]["sha"]

    # Step 2
    commit = github_request(
        "GET",
        f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/git/commits/{latest_commit_sha}"
    )

    base_tree_sha = commit["tree"]["sha"]

    # Step 3
    tree = []

    for file in files:
        tree.append({
            "path": CONTENT_PATH + file["filename"],
            "mode": "100644",
            "type": "blob",
            "content": file["content"]
        })

    tree_result = github_request(
        "POST",
        f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/git/trees",
        json={
            "base_tree": base_tree_sha,
            "tree": tree
        }
    )

    new_tree_sha = tree_result["sha"]

    # Step 4
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    commit_result = github_request(
        "POST",
        f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/git/commits",
        json={
            "message": f"Update {len(files)} product posts - {timestamp}",
            "tree": new_tree_sha,
            "parents": [latest_commit_sha]
        }
    )

    new_commit_sha = commit_result["sha"]

    # Step 5
    github_request(
        "PATCH",
        f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/git/refs/heads/{BRANCH}",
        json={
            "sha": new_commit_sha,
            "force": False
        }
    )

    print(f"Successfully pushed {len(files)} markdown files in one commit.")


# -----------------------------------
# Main
# -----------------------------------

if __name__ == "__main__":

    print("1. Save markdown locally")
    print("2. Push to GitHub")

    choice = input("> ").strip()

    if choice == "1":
        save_markdown_files()

    elif choice == "2":
        push_to_github()

    else:
        print("Invalid option.")
